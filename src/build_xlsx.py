"""
Gera a planilha pública a partir de `data/series.parquet`.

Estrutura da pasta:
  Leia-me       procedência, abrangência das fontes e a ressalva BIS × SFN
  Dicionário    uma linha por série: nome oficial, código na fonte, unidade, cobertura
  <bloco>       uma aba por bloco temático, datas em linhas e séries em colunas
  dados_longo   o formato canônico inteiro, para quem for reprocessar

A planilha é gerada, nunca editada à mão. Copiada para `docs/` porque o GitHub Pages
publica apenas o conteúdo dessa pasta.

Uso:
    python src/build_xlsx.py
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from comum import CONFIG, DATA, DOCS, carrega_blocos, carrega_catalogo

ARQUIVO = "monitor_endividamento.xlsx"

AZUL = "164194"

LEIA_ME = [
    ("Monitor de Endividamento", ""),
    ("Desenvolvido por Kleber Pacheco de Castro", ""),
    ("", ""),
    (
        "Painel técnico de acompanhamento de dados públicos. Não constitui posição "
        "institucional de nenhuma entidade nem recomendação de investimento.",
        "",
    ),
    ("", ""),
    ("Abrangência das fontes", ""),
    (
        "BCB/SGS",
        "Cobre exclusivamente operações do Sistema Financeiro Nacional. Não inclui dívida "
        "com o comércio, com fintechs não reguladas, nem captação em mercado de capitais "
        "ou no exterior.",
    ),
    (
        "BIS (via FRED)",
        "Cobre crédito ao setor de todas as fontes — bancos domésticos, mercado de capitais "
        "e credores externos. Os níveis são estruturalmente mais altos que os do SFN e NÃO "
        "são comparáveis com as abas do bloco Brasil.",
    ),
    ("", ""),
    (
        "Tratamento dos dados",
        "Nenhum valor é interpolado, arredondado ou estimado. Observação sem valor "
        "divulgado na fonte é omitida. Cada série mantém a unidade original da fonte.",
    ),
    (
        "Séries calculadas",
        "As séries terminadas em _real não vêm da fonte: são o saldo nominal deflacionado "
        "pelo IPCA (SGS 433), a preços do mês mais recente do índice, conforme a regra em "
        "config/derivadas.yaml. A unidade de cada uma diz qual é o mês-base. As séries "
        "nominais correspondentes seguem na planilha, sem alteração.",
    ),
]


def _catalogo_completo() -> dict[str, dict]:
    # `derivadas.yaml` entra aqui para que as séries calculadas apareçam na aba do bloco
    # e no Dicionário, ao lado das coletadas.
    series = {}
    for arquivo in ("series_bcb.yaml", "series_fred.yaml", "derivadas.yaml"):
        for serie in carrega_catalogo(arquivo)["series"]:
            series[serie["serie_id"]] = serie
    return series


def _largura(ws, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _cabecalho(ws) -> None:
    """Deixa a primeira linha em negrito azul e congela o cabeçalho."""
    for celula in ws[1]:
        celula.font = Font(bold=True, color=AZUL)
    ws.freeze_panes = "A2"


def _nomes_oficiais() -> dict[str, str]:
    """
    Nome com que a série é publicada na fonte, colhido por `validate_series.py`.

    Vem de `config/_validacao.json`, que guarda o título retornado pela própria API —
    é o nome oficial de fato, não a expectativa registrada no YAML.
    """
    caminho = CONFIG / "_validacao.json"
    if not caminho.exists():
        return {}
    try:
        registros = json.loads(caminho.read_text(encoding="utf-8"))
    except ValueError:
        return {}
    return {r["serie_id"]: r.get("nome_oficial") or "" for r in registros}


def monta_dicionario(catalogo: dict[str, dict], manifesto: list[dict]) -> pd.DataFrame:
    """Uma linha por série coletada: identificação, procedência e cobertura."""
    oficiais = _nomes_oficiais()
    linhas = []
    for entrada in manifesto:
        cfg = catalogo.get(entrada["serie_id"], {})
        linhas.append(
            {
                "serie_id": entrada["serie_id"],
                "nome_oficial": oficiais.get(entrada["serie_id"])
                or cfg.get("descricao_esperada", ""),
                "fonte": entrada["fonte"],
                "codigo_na_fonte": entrada["codigo_fonte"],
                "unidade": entrada["unidade"],
                "periodicidade": entrada["periodicidade"],
                "primeira_observacao": entrada.get("inicio"),
                "ultima_observacao": entrada.get("ultima_data"),
            }
        )
    return pd.DataFrame(linhas)


def para_largo(df: pd.DataFrame, series_ids: list[str]) -> pd.DataFrame:
    """Pivota o formato longo para data × série, preservando lacunas como vazio."""
    recorte = df[df["serie_id"].isin(series_ids)]
    largo = recorte.pivot_table(index="data", columns="serie_id", values="valor", aggfunc="first")
    ordem = [s for s in series_ids if s in largo.columns]
    largo = largo[ordem].sort_index()
    largo.index = largo.index.date
    largo.index.name = "data"
    return largo.reset_index()


def main() -> int:
    caminho_parquet = DATA / "series.parquet"
    if not caminho_parquet.exists():
        print("data/series.parquet não existe. Rode `python src/build_dataset.py` antes.")
        return 1

    df = pd.read_parquet(caminho_parquet)
    manifesto = json.loads((DATA / "manifest.json").read_text(encoding="utf-8"))
    catalogo = _catalogo_completo()
    blocos = carrega_blocos()["blocos"]

    destino = DATA / ARQUIVO
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        leia_me = pd.DataFrame(LEIA_ME, columns=["", ""])
        leia_me.loc[len(leia_me)] = ["Gerado em (UTC)", manifesto["gerado_em"]]
        leia_me.to_excel(writer, sheet_name="Leia-me", index=False, header=False)
        ws = writer.sheets["Leia-me"]
        _largura(ws, [28, 110])
        ws["A1"].font = Font(bold=True, color=AZUL, size=14)
        for linha in ws.iter_rows(min_col=2, max_col=2):
            for celula in linha:
                celula.alignment = Alignment(wrap_text=True, vertical="top")

        dicionario = monta_dicionario(catalogo, manifesto["series"])
        dicionario.to_excel(writer, sheet_name="Dicionário", index=False)
        _largura(writer.sheets["Dicionário"], [38, 96, 12, 16, 14, 14, 20, 20])
        _cabecalho(writer.sheets["Dicionário"])

        for bloco in blocos:
            ids = [s["serie_id"] for s in catalogo.values() if s.get("bloco") == bloco["id"]]
            ids = [i for i in ids if i in set(df["serie_id"])]
            if not ids:
                continue
            # `aba` explícita em blocos.yaml; sem ela, corta no limite do Excel.
            aba = (bloco.get("aba") or bloco["titulo"])[:31]
            para_largo(df, ids).to_excel(writer, sheet_name=aba, index=False)
            _largura(writer.sheets[aba], [12] + [26] * len(ids))
            _cabecalho(writer.sheets[aba])

        longo = df.copy()
        longo["data"] = longo["data"].dt.date
        longo.to_excel(writer, sheet_name="dados_longo", index=False)
        _largura(writer.sheets["dados_longo"], [36, 12, 14, 12, 16, 14, 14])
        _cabecalho(writer.sheets["dados_longo"])

    DOCS.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(destino, DOCS / ARQUIVO)

    print(f"Escrito: data/{ARQUIVO} e docs/{ARQUIVO} ({destino.stat().st_size // 1024} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
